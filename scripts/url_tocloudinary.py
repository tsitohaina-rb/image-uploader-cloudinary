#!/usr/bin/env python3
"""
URL to Cloudinary Converter - Enhanced Version

This script can handle multiple input types with enterprise-grade features:
1. CSV/XLS/XLSX files with CDN image URLs in columns H-Q (media columns)
2. Single image URL processing

Features:
- Multiprocessing for maximum performance
- Comprehensive caching and resume functionality
- Cross-platform file locking
- Exponential backoff retry logic
- Progress tracking and statistics
- Connection pooling optimization
- Comprehensive logging

Usage:
    # For files
    python scripts/url_tocloudinary.py input_file.csv [output_file.xlsx]
    python scripts/url_tocloudinary.py input_file.xlsx [output_file.xlsx]
    
    # For single URL
    python scripts/url_tocloudinary.py --url "https://example.com/image.jpg"
    
    # Advanced options
    python scripts/url_tocloudinary.py input.csv --workers 8 --folder "my_folder" --force-rescan

Requirements:
    - requests, pandas, openpyxl, xlrd, cloudinary, PIL/Pillow
"""

import os
import sys
import csv
import json
import hashlib
import logging
import requests
import tempfile
import cloudinary.uploader
import glob
import multiprocessing
import platform
import socket
import ssl
import time
import urllib.parse
import re
import argparse
import unicodedata
import shutil
import shutil
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor, as_completed
from multiprocessing import Manager, Lock as MPLock
from typing import Optional
import io
from functools import wraps
from PIL import Image, ImageOps
from tqdm import tqdm
from datetime import datetime
import pandas as pd
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# Windows console encoding setup for emoji/Unicode support
if platform.system() == "Windows":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except (AttributeError, OSError):
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'ignore')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'ignore')

# Cross-platform emoji/symbol compatibility
IS_WINDOWS = platform.system() == "Windows"
SYMBOLS = {
    'check': 'OK' if IS_WINDOWS else '✓',
    'cross': 'X' if IS_WINDOWS else '❌',
    'warning': '[!]' if IS_WINDOWS else '⚠️',
    'upload': '[UP]' if IS_WINDOWS else '⬆️',
    'success': '[OK]' if IS_WINDOWS else '✅',
    'failed': '[FAIL]' if IS_WINDOWS else '❌',
    'skip': '[SKIP]' if IS_WINDOWS else '⏭️',
    'compress': '[ZIP]' if IS_WINDOWS else '🗜️',
    'folder': '[DIR]' if IS_WINDOWS else '📁',
    'cache': '[CACHE]' if IS_WINDOWS else '💾',
    'resume': '[RESUME]' if IS_WINDOWS else '🔄',
    'new': '[NEW]' if IS_WINDOWS else '🆕',
}

# Cross-platform file locking
try:
    import fcntl  # Unix/Linux/macOS
    HAS_FCNTL = True
except ImportError:
    try:
        import msvcrt  # Windows
        HAS_FCNTL = False
    except ImportError:
        HAS_FCNTL = None

def lock_file(file_handle):
    """Cross-platform file locking"""
    if HAS_FCNTL is True:
        fcntl.flock(file_handle.fileno(), fcntl.LOCK_EX)
    elif HAS_FCNTL is False:
        msvcrt.locking(file_handle.fileno(), msvcrt.LK_LOCK, 1)

def unlock_file(file_handle):
    """Cross-platform file unlocking"""
    if HAS_FCNTL is True:
        fcntl.flock(file_handle.fileno(), fcntl.LOCK_UN)
    elif HAS_FCNTL is False:
        msvcrt.locking(file_handle.fileno(), msvcrt.LK_UNLCK, 1)

# Add parent directory to path to import config
sys.path.append(str(Path(__file__).parent.parent))
try:
    from config import USE_FILENAME, UNIQUE_FILENAME, CLOUDINARY_FOLDER
except ImportError:
    USE_FILENAME = True
    UNIQUE_FILENAME = False
    CLOUDINARY_FOLDER = "url_upload"

# Setup directory structure
DATA_DIR = Path('data')
CACHE_DIR = DATA_DIR / 'cache'
LOG_DIR = DATA_DIR / 'log' 
OUTPUT_DIR = DATA_DIR / 'output'
TEMP_DIR = DATA_DIR / 'temp'

for directory in [CACHE_DIR, LOG_DIR, OUTPUT_DIR, TEMP_DIR]:
    directory.mkdir(parents=True, exist_ok=True)

# Set socket timeout globally
socket.setdefaulttimeout(60)

# SSL Context configuration for better stability
ssl_context = ssl.create_default_context()
ssl_context.check_hostname = False
ssl_context.verify_mode = ssl.CERT_NONE

# Media columns mapping (H to Q in Excel = columns 7-16 in 0-indexed Python)
MEDIA_COLUMNS = [
    'Image principale', 'image secondaire', 'Image 3', 'Image 4', 'Image 5',
    'Image 6', 'Image 7', 'Image 8', 'Image 9', 'Image_10'
]

# Global variables for multiprocessing
progress_lock = None
progress_counter = None
error_log = None

def get_process_id():
    """Get a short process identifier for logging"""
    return f"P{os.getpid() % 1000:03d}"

# Retry decorator with exponential backoff
def retry_with_backoff(max_retries=3, backoff_factor=2, exceptions=(Exception,)):
    """Decorator to retry functions with exponential backoff on specified exceptions."""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            process_id = get_process_id()
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except exceptions as e:
                    if attempt == max_retries - 1:
                        raise e
                    
                    error_type = type(e).__name__
                    wait_time = backoff_factor ** attempt
                    
                    logging.warning(f"  [Process {process_id}] Attempt {attempt + 1} failed with {error_type}: {str(e)}")
                    logging.info(f"  [Process {process_id}] Retrying in {wait_time:.2f} seconds...")
                    time.sleep(wait_time)
            return None
        return wrapper
    return decorator

def create_robust_session():
    """Create a requests session with robust configuration"""
    session = requests.Session()
    
    retry_strategy = Retry(
        total=3,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET", "OPTIONS"],
        backoff_factor=1
    )
    
    adapter = HTTPAdapter(
        pool_connections=20,
        pool_maxsize=50,
        max_retries=retry_strategy
    )
    
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    
    return session

def setup_logging(base_name, folder_name=None):
    """Setup logging with timestamp and folder-specific log file"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Use folder name if provided, otherwise use base_name
    if folder_name:
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', f"url_to_cloudinary_{folder_name}").strip()
    else:
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', base_name).strip()
    
    log_filename = LOG_DIR / f'{safe_name}_{timestamp}.log'
    
    file_handler = logging.FileHandler(str(log_filename), encoding='utf-8')
    console_handler = logging.StreamHandler()
    
    if platform.system() == 'Windows':
        try:
            console_handler.stream = open(console_handler.stream.fileno(), mode='w', encoding='utf-8', buffering=1)
        except:
            pass
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[file_handler, console_handler]
    )
    return str(log_filename)

class URLUploadCache:
    """Manages the cache of uploaded URLs to support resume functionality"""
    
    def __init__(self, input_file_path: str, folder_name: Optional[str] = None):
        """Initialize cache for a specific input file and optional folder"""
        self.input_file_path = input_file_path
        
        # Create cache file name based on folder or input file
        if folder_name:
            safe_name = re.sub(r'[<>:"/\\|?*]', '_', folder_name).strip()
        else:
            input_name = Path(input_file_path).stem
            safe_name = re.sub(r'[<>:"/\\|?*]', '_', input_name).strip()
        
        # Look for existing cache files first
        cache_pattern = str(CACHE_DIR / f'url_upload_cache_{safe_name}_*.json')
        existing_caches = glob.glob(cache_pattern)
        
        if existing_caches:
            self.cache_file = Path(max(existing_caches, key=os.path.getctime))
        else:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.cache_file = CACHE_DIR / f'url_upload_cache_{safe_name}_{timestamp}.json'
        
        from threading import Lock
        self.lock = Lock()
        self.cache = self._load_cache()
        
        # Add cache throttling to prevent too frequent writes
        self.last_save_time = 0
        self.save_interval = 2.0  # Save at most every 2 seconds
        self.pending_saves = 0
        
        # Add cache throttling to prevent too frequent writes
        self.last_save_time = 0
        self.save_interval = 2.0  # Save at most every 2 seconds
        self.pending_saves = 0
    
    def _load_cache(self) -> dict:
        """Load cache with corruption recovery"""
        default_cache = {
            'input_file': self.input_file_path,
            'last_run': '',
            'successful_uploads': {},
            'failed_uploads': {}
        }
        
        if not self.cache_file.exists():
            return default_cache
        
        # Try to load cache file with corruption recovery
        for attempt_file in [self.cache_file, self.cache_file.with_suffix('.backup')]:
            if attempt_file.exists():
                try:
                    with open(attempt_file, 'r', encoding='utf-8') as f:
                        lock_file(f)
                        try:
                            data = json.load(f)
                            logging.info(f"Loaded cache from {attempt_file.name} with {len(data.get('successful_uploads', {}))} successful entries")
                            return data
                        finally:
                            unlock_file(f)
                except json.JSONDecodeError as e:
                    logging.warning(f"Cache file {attempt_file.name} corrupted (JSON error): {e}")
                    
                    # Only show interactive prompt during main process initialization, not during worker processes
                    if multiprocessing.current_process().name == 'MainProcess':
                        print(f"\n{SYMBOLS['warning']} CACHE CORRUPTION DETECTED!")
                        print(f"   File: {attempt_file}")
                        print(f"   Error: {e}")
                        print(f"   This contains your upload progress - don't lose it!")
                        print(f"\n{SYMBOLS['warning']} PROCESSING PAUSED")
                        print(f"   Please fix the cache file manually and then:")
                        print(f"   - Type 'continue' or 'yes' to retry loading the cache")
                        print(f"   - Type 'skip' to start with empty cache (LOSES PROGRESS!)")
                        
                        while True:
                            user_input = input(f"\n   What would you like to do? ").lower().strip()
                            if user_input in ['continue', 'yes', 'c', 'y']:
                                print(f"   {SYMBOLS['upload']} Retrying cache load...")
                                # Retry loading the same file
                                try:
                                    with open(attempt_file, 'r', encoding='utf-8') as f:
                                        lock_file(f)
                                        try:
                                            data = json.load(f)
                                            logging.info(f"Successfully loaded fixed cache from {attempt_file.name} with {len(data.get('successful_uploads', {}))} entries")
                                            print(f"   {SYMBOLS['success']} Cache loaded successfully!")
                                            return data
                                        finally:
                                            unlock_file(f)
                                except Exception as retry_e:
                                    print(f"   {SYMBOLS['failed']} Still corrupted: {retry_e}")
                                    print(f"   Please fix the file and try again.")
                                    continue
                            elif user_input in ['skip', 's']:
                                print(f"   {SYMBOLS['warning']} Skipping corrupted cache - starting fresh (progress lost)")
                                # Move corrupted file as backup
                                corrupted_file = attempt_file.with_suffix('.corrupted')
                                try:
                                    shutil.move(str(attempt_file), str(corrupted_file))
                                    logging.info(f"Moved corrupted cache to: {corrupted_file}")
                                except:
                                    pass
                                break
                            else:
                                print(f"   Please type 'continue'/'yes' to retry or 'skip' to start fresh")
                    else:
                        # In worker process, just skip corrupted cache silently
                        logging.warning(f"Worker process detected corrupted cache, skipping: {attempt_file.name}")
                        corrupted_file = attempt_file.with_suffix('.corrupted')
                        try:
                            shutil.move(str(attempt_file), str(corrupted_file))
                            logging.info(f"Moved corrupted cache to: {corrupted_file}")
                        except:
                            pass
                except Exception as e:
                    logging.warning(f"Could not load cache file {attempt_file.name}: {e}")
        
        logging.warning("All cache files corrupted or missing, starting with empty cache")
        return default_cache
    
    def _save_cache(self):
        """Save cache with atomic write to prevent corruption"""
        try:
            # Create cache directory if needed
            self.cache_file.parent.mkdir(parents=True, exist_ok=True)
            
            # Use atomic write with temporary file
            temp_file = self.cache_file.with_suffix('.tmp')
            
            # First, try to merge with existing cache
            existing_cache = {}
            if self.cache_file.exists():
                try:
                    with open(self.cache_file, 'r', encoding='utf-8') as f:
                        lock_file(f)
                        try:
                            existing_cache = json.load(f)
                        finally:
                            unlock_file(f)
                except (json.JSONDecodeError, OSError) as e:
                    logging.warning(f"Cache file corrupted, starting fresh: {e}")
                    # Move corrupted file as backup
                    backup_file = self.cache_file.with_suffix('.corrupted')
                    try:
                        shutil.move(str(self.cache_file), str(backup_file))
                        logging.info(f"Corrupted cache moved to: {backup_file}")
                    except:
                        pass
            
            # Merge caches
            merged_cache = {
                'input_file': self.input_file_path,
                'last_run': datetime.now().isoformat(),
                'successful_uploads': {},
                'failed_uploads': {}
            }
            
            # Start with existing cache data
            if 'successful_uploads' in existing_cache:
                merged_cache['successful_uploads'].update(existing_cache['successful_uploads'])
            if 'failed_uploads' in existing_cache:
                merged_cache['failed_uploads'].update(existing_cache['failed_uploads'])
            
            # Add current cache data
            merged_cache['successful_uploads'].update(self.cache['successful_uploads'])
            merged_cache['failed_uploads'].update(self.cache['failed_uploads'])
            
            # Remove failed entries that are now successful
            for key in merged_cache['successful_uploads']:
                merged_cache['failed_uploads'].pop(key, None)
            
            # Write to temporary file first (atomic operation)
            with open(temp_file, 'w', encoding='utf-8') as f:
                lock_file(f)
                try:
                    json.dump(merged_cache, f, indent=2, ensure_ascii=False)
                    f.flush()
                    try:
                        os.fsync(f.fileno())  # Force write to disk
                    except (OSError, AttributeError):
                        pass  # fsync not available on all systems
                finally:
                    unlock_file(f)
            
            # Atomic rename (this is the critical atomic operation)
            if temp_file.exists():
                # On Windows, we need to remove the target file first
                if platform.system() == 'Windows' and self.cache_file.exists():
                    self.cache_file.unlink()
                temp_file.rename(self.cache_file)
            
            # Update our local cache with merged data
            self.cache = merged_cache
            
        except Exception as e:
            logging.warning(f"Could not save cache file: {e}")
            # Clean up temp file if it exists
            temp_file = self.cache_file.with_suffix('.tmp')
            if temp_file.exists():
                try:
                    temp_file.unlink()
                except:
                    pass
    
    def _save_cache_throttled(self, force=False):
        """Save cache with throttling to prevent too frequent writes"""
        current_time = time.time()
        
        if force or (current_time - self.last_save_time) >= self.save_interval:
            self._save_cache()
            self.last_save_time = current_time
            self.pending_saves = 0
        else:
            self.pending_saves += 1
    
    def get_url_key(self, url, row_index, column_name):
        """Generate a unique key for URL cache"""
        url_hash = hashlib.md5(url.encode()).hexdigest()[:8]
        return f"{row_index}_{column_name}_{url_hash}"
    
    def is_uploaded(self, url, row_index, column_name) -> bool:
        """Check if a URL was successfully uploaded"""
        key = self.get_url_key(url, row_index, column_name)
        with self.lock:
            return key in self.cache['successful_uploads']
    
    def mark_uploaded(self, url, row_index, column_name, result: dict):
        """Mark a URL as successfully uploaded"""
        key = self.get_url_key(url, row_index, column_name)
        with self.lock:
            self.cache['last_run'] = datetime.now().isoformat()
            self.cache['successful_uploads'][key] = {
                'timestamp': datetime.now().isoformat(),
                'original_url': url,
                'cloudinary_url': result['cloudinary_url'],
                'public_id': result.get('public_id', ''),
                'row': row_index,
                'column': column_name
            }
            if key in self.cache['failed_uploads']:
                del self.cache['failed_uploads'][key]
            self._save_cache_throttled()
    
    def mark_failed(self, url, row_index, column_name, error: str):
        """Mark a URL as failed upload"""
        key = self.get_url_key(url, row_index, column_name)
        with self.lock:
            self.cache['last_run'] = datetime.now().isoformat()
            self.cache['failed_uploads'][key] = {
                'timestamp': datetime.now().isoformat(),
                'original_url': url,
                'error': error,
                'row': row_index,
                'column': column_name
            }
            self._save_cache_throttled()
    
    def get_cached_result(self, url, row_index, column_name):
        """Get cached result for a URL"""
        key = self.get_url_key(url, row_index, column_name)
        with self.lock:
            if key in self.cache['successful_uploads']:
                cached = self.cache['successful_uploads'][key]
                return {
                    'original_url': cached['original_url'],
                    'cloudinary_url': cached['cloudinary_url'],
                    'status': 'success',
                    'public_id': cached.get('public_id', ''),
                    'row': str(cached['row']),
                    'column': cached['column'],
                    'cached': True
                }
        return None
    
    def get_stats(self) -> dict:
        """Get cache statistics"""
        return {
            'successful': len(self.cache['successful_uploads']),
            'failed': len(self.cache['failed_uploads']),
            'last_run': self.cache.get('last_run', '')
        }
    
    def finalize_cache(self):
        """Force save any pending cache changes"""
        if self.pending_saves > 0:
            self._save_cache_throttled(force=True)

def sanitize_filename_for_cloudinary(filename):
    """Sanitize filename for Cloudinary upload"""
    if not filename:
        return 'untitled'
    
    name = os.path.splitext(filename)[0]
    name = re.sub(r'[\s\-\.]+', '_', name)
    name = re.sub(r'[^a-zA-Z0-9_\-]', '', name)
    name = name.strip('_-')
    
    if not name:
        return 'untitled'
    
    return name[:200]

def get_url_filename(url):
    """Extract filename from URL"""
    try:
        parsed = urllib.parse.urlparse(url)
        filename = os.path.basename(parsed.path)
        filename = filename.split('?')[0]
        return filename if filename else 'image'
    except Exception:
        return 'image'

def is_valid_image_url(url):
    """Check if URL is a valid image URL"""
    if not url or not isinstance(url, str):
        return False
    
    url = url.strip()
    if not url or url == 'nan':
        return False
        
    # Must start with http/https
    if not url.startswith(('http://', 'https://')):
        return False
    
    # Skip header placeholders and error messages
    if url.startswith(('media_', 'Image', 'DOWNLOAD_FAILED', 'UPLOAD_FAILED', 'PROCESSING_FAILED', 'TASK_FAILED')):
        return False
    
    # Basic URL structure validation
    try:
        parsed = urllib.parse.urlparse(url)
        return bool(parsed.netloc and parsed.scheme in ['http', 'https'])
    except Exception:
        return False

@retry_with_backoff(max_retries=3, exceptions=(requests.RequestException,))
def download_image(url, session, timeout=30):
    """Download image from URL using provided session"""
    if not url or not url.strip():
        return None, "Empty URL"
    
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        response = session.get(url, headers=headers, timeout=timeout)
        response.raise_for_status()
        
        content_type = response.headers.get('content-type', '').lower()
        if not content_type.startswith('image/'):
            return None, f"Not an image: content-type is {content_type}"
        
        return response.content, None
        
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 404:
            return None, "404 Not Found"
        return None, f"HTTP Error {e.response.status_code}: {e}"
    except requests.exceptions.RequestException as e:
        return None, f"Request Error: {e}"
    except Exception as e:
        return None, f"Unexpected Error: {e}"

def compress_image_for_cloudinary(image_data, filename, max_size_mb=20, quality=85):
    """Compress image data to reduce file size for Cloudinary upload"""
    try:
        original_size = len(image_data)
        original_size_mb = original_size / (1024 * 1024)
        
        if original_size_mb <= max_size_mb:
            return image_data, False, original_size_mb, 1.0
        
        image = Image.open(io.BytesIO(image_data))
        
        if image.mode in ('RGBA', 'LA', 'P'):
            background = Image.new('RGB', image.size, (255, 255, 255))
            if image.mode == 'P':
                image = image.convert('RGBA')
            background.paste(image, mask=image.split()[-1] if image.mode in ('RGBA', 'LA') else None)
            image = background
        
        image = ImageOps.exif_transpose(image)
        
        output_buffer = io.BytesIO()
        image.save(output_buffer, format='JPEG', quality=quality, optimize=True)
        compressed_data = output_buffer.getvalue()
        
        final_size = len(compressed_data)
        final_size_mb = final_size / (1024 * 1024)
        compression_ratio = final_size / original_size
        
        logging.info(f"Compressed {filename}: {original_size_mb:.2f}MB → {final_size_mb:.2f}MB "
                    f"(ratio: {compression_ratio:.2f}, quality: {quality})")
        
        return compressed_data, True, final_size_mb, compression_ratio
        
    except Exception as e:
        logging.warning(f"Compression failed for {filename}: {e}")
        original_size_mb = len(image_data) / (1024 * 1024)
        return image_data, False, original_size_mb, 1.0

@retry_with_backoff(max_retries=3, exceptions=(Exception,))
def upload_to_cloudinary(image_data, filename, folder_name=None):
    """Upload image data to Cloudinary"""
    try:
        safe_filename = sanitize_filename_for_cloudinary(filename)
        
        processed_data, was_compressed, size_mb, ratio = compress_image_for_cloudinary(
            image_data, filename
        )
        
        upload_kwargs = {
            'folder': folder_name or CLOUDINARY_FOLDER,
            'public_id': safe_filename,
            'resource_type': 'image',
            'use_filename': USE_FILENAME,
            'unique_filename': UNIQUE_FILENAME,
            'overwrite': True,
            'timeout': 120
        }
        
        response = cloudinary.uploader.upload(
            io.BytesIO(processed_data),
            **upload_kwargs
        )
        
        return {
            'cloudinary_url': response['secure_url'],
            'public_id': response.get('public_id', ''),
            'status': 'success',
            'compressed': was_compressed,
            'size_mb': size_mb
        }
        
    except Exception as e:
        logging.error(f"Cloudinary upload failed for {filename}: {e}")
        return {
            'cloudinary_url': 'UPLOAD_FAILED',
            'public_id': '',
            'status': 'failed',
            'error': str(e)
        }

def update_progress(status):
    """Thread-safe progress update"""
    global progress_lock, progress_counter
    if progress_lock and progress_counter:
        with progress_lock:
            progress_counter[status] = progress_counter.get(status, 0) + 1

def _url_upload_worker(task_data):
    """
    Worker function for multiprocessing URL uploads.
    Each process creates its own session and cache.
    """
    url, row_index, column_name, folder_name, input_file_path = task_data
    
    process_id = get_process_id()
    
    # Validate URL before processing
    if not is_valid_image_url(url):
        return {
            'original_url': url,
            'cloudinary_url': url,  # Keep original URL instead of INVALID_URL
            'status': 'invalid_url',
            'error': 'Invalid URL format or header placeholder',
            'row': str(row_index),
            'column': column_name,
            'failed': 'true'  # Flag to indicate this should be colored red
        }
    
    # Each process needs its own session and cache (simplified for workers)
    session = create_robust_session()
    
    # For worker processes, create a simple cache instance without interactive prompts
    try:
        cache = URLUploadCache(input_file_path)
    except Exception as cache_error:
        # If cache fails in worker, continue without cache
        logging.warning(f"[{process_id}] Cache unavailable in worker: {cache_error}")
        cache = None
    
    # Check cache first (if available)
    cached_result = None
    if cache:
        cached_result = cache.get_cached_result(url, row_index, column_name)
    
    if cached_result:
        update_progress('skipped')
        logging.info(f"[{process_id}] {SYMBOLS['skip']} Cached: Row {row_index}, {column_name}")
        return cached_result
    
    # Download image
    logging.info(f"[{process_id}] {SYMBOLS['upload']} Processing Row {row_index}, {column_name}: {url}")
    
    try:
        image_data, download_error = download_image(url, session)
        
        if download_error:
            if cache:
                cache.mark_failed(url, row_index, column_name, download_error)
            update_progress('failed')
            
            status = 'download_failed'
            if '404' in download_error:
                status = 'not_found'
            
            return {
                'original_url': url,
                'cloudinary_url': url,  # Keep original URL instead of DOWNLOAD_FAILED
                'status': status,
                'error': download_error,
                'row': str(row_index),
                'column': column_name,
                'failed': 'true'  # Flag to indicate this should be colored red
            }
        
        # Generate filename
        filename = get_url_filename(url)
        if not filename:
            filename = f"image_{row_index}_{column_name}"
        
        # Upload to Cloudinary
        result = upload_to_cloudinary(image_data, filename, folder_name)
        
        if result['status'] == 'success':
            if cache:
                cache.mark_uploaded(url, row_index, column_name, result)
            update_progress('uploaded')
            logging.info(f"[{process_id}] {SYMBOLS['success']} Uploaded: {result['cloudinary_url']}")
        else:
            if cache:
                cache.mark_failed(url, row_index, column_name, result.get('error', 'Unknown error'))
            update_progress('failed')
            # Keep original URL for failed uploads
            result['cloudinary_url'] = url
            result['failed'] = 'true'  # String instead of boolean for consistency
        
        result['original_url'] = url
        result['row'] = str(row_index)
        result['column'] = column_name
        
        return result
        
    except Exception as e:
        error_msg = f"Processing failed: {e}"
        if cache:
            cache.mark_failed(url, row_index, column_name, error_msg)
        update_progress('failed')
        
        return {
            'original_url': url,
            'cloudinary_url': url,  # Keep original URL instead of PROCESSING_FAILED
            'status': 'processing_failed',
            'error': error_msg,
            'row': str(row_index),
            'column': column_name,
            'failed': 'true'  # Flag to indicate this should be colored red
        }
    finally:
        session.close()

def read_input_file(file_path):
    """Read input file (CSV, XLS, or XLSX) and return DataFrame"""
    file_ext = Path(file_path).suffix.lower()
    
    try:
        if file_ext == '.csv':
            df = pd.read_csv(file_path)
        elif file_ext == '.xls':
            df = pd.read_excel(file_path, engine='xlrd')
        elif file_ext == '.xlsx':
            df = pd.read_excel(file_path, engine='openpyxl')
        else:
            raise ValueError(f"Unsupported file format: {file_ext}. Supported: .csv, .xls, .xlsx")
        
        logging.info(f"File loaded: {len(df)} rows, {len(df.columns)} columns")
        return df
        
    except Exception as e:
        logging.error(f"Failed to read file {file_path}: {e}")
        raise

def process_file_to_cloudinary(input_file_path, output_xlsx_path=None, max_workers=None, folder_name=None, force_rescan=False):
    """Main function to process file and convert URLs to Cloudinary URLs"""
    
    # Generate output filename if not provided
    if not output_xlsx_path:
        # Use folder name if provided, otherwise use input file name
        if folder_name:
            output_name = folder_name
        else:
            output_name = Path(input_file_path).stem
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_xlsx_path = OUTPUT_DIR / f"{output_name}_cloudinary_{timestamp}.xlsx"
    
    # Setup logging with folder or input name
    input_name = Path(input_file_path).stem
    log_file = setup_logging(f"url_to_cloudinary_{input_name}", folder_name)
    
    print(f"{SYMBOLS['upload']} Starting URL to Cloudinary conversion")
    print(f"   Input file: {input_file_path}")
    print(f"   Output file: {output_xlsx_path}")
    print(f"   Log file: {log_file}")
    print(f"   Cloudinary folder: {folder_name or CLOUDINARY_FOLDER}")
    
    # Read input file
    df = read_input_file(input_file_path)
    
    # Find media columns (H-Q = columns 7-16)
    media_column_indices = list(range(7, min(17, len(df.columns))))
    media_column_names = [df.columns[idx] for idx in media_column_indices]
    
    print(f"\n{SYMBOLS['folder']} Media columns found: {media_column_names}")
    
    # Initialize cache
    cache = URLUploadCache(input_file_path, folder_name)
    cache_stats = cache.get_stats()
    
    if cache_stats['successful'] > 0 and not force_rescan:
        print(f"\n{SYMBOLS['resume']} RESUMING: Found {cache_stats['successful']} previously uploaded URLs in cache")
        print(f"   Last run: {cache_stats['last_run']}")
        print(f"   Use --force-rescan to ignore cache")
    else:
        if force_rescan:
            print(f"\n{SYMBOLS['new']} FORCE RESCAN: Ignoring cache")
        else:
            print(f"\n{SYMBOLS['new']} STARTING: New upload session")
    
    # Collect URLs to process (skip header row - start from row 1, which is index 1)
    url_tasks = []
    for row_index, row in df.iterrows():
        # Skip the header row (row 0 in 0-indexed, which is row 1 in Excel)
        if row_index == 0:
            continue
            
        for col_idx, col_name in enumerate(media_column_names):
            if col_idx + 7 < len(row):
                cell_value = row.iloc[col_idx + 7] if col_idx + 7 < len(row) else None
                url = str(cell_value).strip() if cell_value is not None and pd.notna(cell_value) else ""
                
                # Only process valid HTTP/HTTPS URLs, skip header placeholders and failed URLs
                if (url and url != 'nan' and 
                    url.startswith(('http://', 'https://')) and 
                    not url.startswith(('DOWNLOAD_FAILED', 'UPLOAD_FAILED', 'PROCESSING_FAILED', 'TASK_FAILED')) and
                    not url.startswith(('media_', 'Image'))):
                    
                    # Skip if cached (unless force rescan)
                    if not force_rescan and cache.is_uploaded(url, row_index, col_name):
                        continue
                    url_tasks.append((url, row_index, col_name, folder_name, input_file_path))
    
    total_urls = len(url_tasks)
    cached_count = cache_stats['successful'] if not force_rescan else 0
    
    print(f"\n{SYMBOLS['upload']} URLs to process: {total_urls}")
    if cached_count > 0:
        print(f"   {SYMBOLS['cache']} Cached (will skip): {cached_count}")
        print(f"   {SYMBOLS['upload']} New to upload: {total_urls}")
    
    if total_urls == 0:
        print(f"\n{SYMBOLS['success']} All URLs already processed! Loading cached results...")
        
        # Load cached results into DataFrame (skip header row)
        for row_index, row in df.iterrows():
            # Skip the header row
            if row_index == 0:
                continue
                
            for col_idx, col_name in enumerate(media_column_names):
                if col_idx + 7 < len(row):
                    cell_value = row.iloc[col_idx + 7]
                    url = str(cell_value).strip() if pd.notna(cell_value) else ""
                    
                    if url and url != 'nan' and url.startswith(('http://', 'https://')):
                        cached_result = cache.get_cached_result(url, row_index, col_name)
                        if cached_result:
                            df.at[row_index, col_name] = cached_result['cloudinary_url']
        
        # Save and return
        save_results_to_xlsx(df, output_xlsx_path, [], [], cache_stats['successful'], 0, 0, {})
        return
    
    # Determine worker count
    if max_workers is None:
        max_workers = min(multiprocessing.cpu_count(), 8)  # Reasonable default
    
    print(f"   {SYMBOLS['upload']} Using {max_workers} worker processes")
    
    # Initialize multiprocessing
    manager = Manager()
    global progress_lock, progress_counter
    progress_lock = manager.Lock()
    progress_counter = manager.dict()
    
    progress_counter['uploaded'] = 0
    progress_counter['failed'] = 0
    progress_counter['skipped'] = 0
    
    # Process URLs with multiprocessing
    processed_results = {}
    failed_urls = []
    not_found_urls = []
    
    print(f"\n{SYMBOLS['upload']} Starting parallel processing...")
    
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        # Submit all tasks
        future_to_task = {
            executor.submit(_url_upload_worker, task): task
            for task in url_tasks
        }
        
        # Process results with progress bar
        with tqdm(total=total_urls, desc="Processing URLs", unit="url") as pbar:
            for future in as_completed(future_to_task):
                task = future_to_task[future]
                url, row_idx, col_name = task[:3]
                
                try:
                    result = future.result()
                    key = (row_idx, col_name)
                    processed_results[key] = result
                    
                    # Track failed URLs
                    if result['status'] == 'download_failed':
                        failed_urls.append(result)
                    elif result['status'] == 'not_found':
                        not_found_urls.append(result)
                    elif result['status'] in ['failed', 'processing_failed']:
                        failed_urls.append(result)
                    
                except Exception as e:
                    logging.error(f"Task failed for {url}: {e}")
                    failed_result = {
                        'original_url': url,
                        'cloudinary_url': url,  # Keep original URL
                        'status': 'task_failed',
                        'error': str(e),
                        'row': str(row_idx),
                        'column': col_name,
                        'failed': 'true'  # Flag for red formatting
                    }
                    failed_urls.append(failed_result)
                    # Also add to processed_results for formatting
                    key = (row_idx, col_name)
                    processed_results[key] = failed_result
                
                pbar.update(1)
    
    # Update DataFrame with results
    for (row_idx, col_name), result in processed_results.items():
        df.at[row_idx, col_name] = result['cloudinary_url']
    
    # Include cached results in DataFrame (skip header row)
    for row_index, row in df.iterrows():
        # Skip the header row
        if row_index == 0:
            continue
            
        for col_idx, col_name in enumerate(media_column_names):
            if col_idx + 7 < len(row):
                cell_value = row.iloc[col_idx + 7]
                url = str(cell_value).strip() if pd.notna(cell_value) else ""
                
                if url and url != 'nan' and url.startswith(('http://', 'https://')) and (row_index, col_name) not in processed_results:
                    cached_result = cache.get_cached_result(url, row_index, col_name)
                    if cached_result:
                        df.at[row_index, col_name] = cached_result['cloudinary_url']
    
    # Calculate final stats
    successful_count = sum(1 for r in processed_results.values() if r['status'] == 'success') + cached_count
    failed_count = len(failed_urls)
    not_found_count = len(not_found_urls)
    
    # Save results
    save_results_to_xlsx(df, output_xlsx_path, failed_urls, not_found_urls, successful_count, failed_count, not_found_count, processed_results)
    
    # Print final summary
    print_final_summary(input_file_path, output_xlsx_path, successful_count, failed_count, not_found_count, not_found_urls)

def save_results_to_xlsx(df, output_path, failed_urls, not_found_urls, successful_count, failed_count, not_found_count, processed_results=None):
    """Save results to XLSX file with multiple sheets and red formatting for failed URLs"""
    try:
        from openpyxl.styles import Font
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Products', index=False)
            
            # Get the workbook and worksheet for styling
            workbook = writer.book
            worksheet = writer.sheets['Products']
            
            # Define red font for failed URLs
            red_font = Font(color='FF0000', bold=True)
            
            # Apply red formatting to failed URLs in the main sheet
            if processed_results:
                # Find media columns
                media_cols = ['Image principale', 'image secondaire', 'Image 3', 'Image 4', 'Image 5',
                             'Image 6', 'Image 7', 'Image 8', 'Image 9', 'Image_10']
                
                for (row_idx, col_name), result in processed_results.items():
                    if result.get('failed') == 'true' and col_name in df.columns:
                        # Find Excel column index
                        col_idx = df.columns.get_loc(col_name) + 1  # +1 for Excel 1-indexing
                        excel_row = row_idx + 2  # +2 for Excel header and 0-indexing
                        
                        cell = worksheet.cell(row=excel_row, column=col_idx)
                        cell.font = red_font
                        
                        # Add comment with error details
                        from openpyxl.comments import Comment
                        error_msg = result.get('error', 'Download failed')
                        comment = Comment(f"FAILED: {error_msg}", "URL_Converter")
                        cell.comment = comment
            
            # Summary sheet
            summary_data = {
                'Metric': [
                    'Successful uploads',
                    'Failed downloads',
                    '404 Not Found',
                    'Upload failures',
                    'Total processed',
                    'Processing date'
                ],
                'Value': [
                    successful_count,
                    failed_count,
                    not_found_count,
                    failed_count - not_found_count,
                    successful_count + failed_count + not_found_count,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Failed URLs sheet
            if failed_urls:
                failed_df = pd.DataFrame(failed_urls)
                failed_df.to_excel(writer, sheet_name='Failed_URLs', index=False)
            
            # 404 Not Found URLs sheet
            if not_found_urls:
                not_found_df = pd.DataFrame(not_found_urls)
                not_found_df.to_excel(writer, sheet_name='404_Not_Found', index=False)
        
        logging.info(f"XLSX file saved: {output_path}")
        
    except Exception as e:
        logging.error(f"Failed to save XLSX file: {e}")
        raise

def print_final_summary(input_file, output_file, successful, failed, not_found, not_found_urls):
    """Print final processing summary"""
    print(f"\n{'='*60}")
    print(f"{SYMBOLS['success']} PROCESSING COMPLETE")
    print(f"{'='*60}")
    print(f"Input file: {input_file}")
    print(f"Output file: {output_file}")
    print(f"Successful uploads: {successful}")
    print(f"Failed downloads: {failed}")
    print(f"404 Not Found: {not_found}")
    print(f"{'='*60}")
    
    if not_found_urls:
        print(f"\n{SYMBOLS['warning']} 404 Not Found URLs ({len(not_found_urls)}):")
        for url_info in not_found_urls[:10]:
            print(f"  Row {url_info['row']}, {url_info['column']}: {url_info['original_url']}")
        if len(not_found_urls) > 10:
            print(f"  ... and {len(not_found_urls) - 10} more (see 404_Not_Found sheet)")

def list_404_urls(input_file_path, output_format='console'):
    """List all rows that contain 404 URLs by actually testing the URLs"""
    print(f"{SYMBOLS['upload']} Analyzing file for 404 URLs: {input_file_path}")
    print(f"   {SYMBOLS['warning']} This will test each URL by making HTTP requests...")
    
    # Read input file
    try:
        df = read_input_file(input_file_path)
    except Exception as e:
        print(f"Error reading file: {e}")
        return
    
    # Find media columns by name (not index)
    media_column_names = ['Image principale', 'image secondaire', 'Image 3', 'Image 4', 'Image 5',
                          'Image 6', 'Image 7', 'Image 8', 'Image 9', 'Image_10']
    
    # Filter to only existing columns
    existing_media_cols = [col for col in media_column_names if col in df.columns]
    
    print(f"\n{SYMBOLS['folder']} Media columns found: {existing_media_cols}")
    
    # Create session for URL testing
    session = create_robust_session()
    # Set timeout in requests, not on session object
    
    # Collect all URLs to test (skip header row)
    url_tests = []
    total_urls = 0
    
    for row_number, (row_index, row) in enumerate(df.iterrows(), start=1):  # start=1 for row numbering
        # Skip the header row (row_index 0)
        if row_index == 0:
            continue
            
        excel_row_number = row_number + 1  # +1 because Excel starts at 1 and we have header
        
        for col_name in existing_media_cols:
            if col_name in row.index:
                cell_value = row[col_name]
                url = str(cell_value).strip() if cell_value is not None and pd.notna(cell_value) else ""
                
                # Only test actual HTTP/HTTPS URLs, skip header placeholders
                if (url and url != 'nan' and 
                    url.startswith(('http://', 'https://')) and
                    not url.startswith(('media_', 'Image'))):
                    url_tests.append({
                        'url': url,
                        'row_number': excel_row_number,
                        'row_index': row_index,
                        'column': col_name
                    })
                    total_urls += 1
    
    print(f"\n{SYMBOLS['upload']} Found {total_urls} URLs to test")
    
    if total_urls == 0:
        print(f"{SYMBOLS['success']} No URLs found to test!")
        return
    
    # Test URLs for 404s
    print(f"{SYMBOLS['upload']} Testing URLs for 404 errors...")
    rows_with_404 = []
    total_404_count = 0
    
    with tqdm(total=total_urls, desc="Testing URLs", unit="url") as pbar:
        for url_test in url_tests:
            url = url_test['url']
            row_number = url_test['row_number']
            column = url_test['column']
            
            try:
                # Test URL with HEAD request (faster than GET)
                response = session.head(url, timeout=10, allow_redirects=True)
                is_404 = response.status_code == 404
                
                # If HEAD fails, try GET (some servers don't support HEAD)
                if response.status_code == 405:  # Method not allowed
                    response = session.get(url, timeout=10, stream=True)
                    is_404 = response.status_code == 404
                    
            except Exception as e:
                # Treat connection errors as potential 404s
                is_404 = True
                error_msg = str(e)
            else:
                error_msg = f"HTTP {response.status_code}" if is_404 else None
            
            if is_404:
                # Find if this row already exists
                existing_row = None
                for row_data in rows_with_404:
                    if row_data['row_number'] == row_number:
                        existing_row = row_data
                        break
                
                if existing_row:
                    existing_row['404_urls'].append({
                        'column': column,
                        'url': url,
                        'error': error_msg
                    })
                else:
                    rows_with_404.append({
                        'row_number': row_number,
                        'row_index': url_test['row_index'],
                        '404_urls': [{
                            'column': column,
                            'url': url,
                            'error': error_msg
                        }]
                    })
                
                total_404_count += 1
            
            pbar.update(1)
    
    # Close session
    session.close()
    
    # Display summary
    print(f"\n{'='*60}")
    print(f"{SYMBOLS['warning']} 404 URL ANALYSIS RESULTS")
    print(f"{'='*60}")
    print(f"Total URLs tested: {total_urls}")
    print(f"Total rows with 404 URLs: {len(rows_with_404)}")
    print(f"Total 404 URLs found: {total_404_count}")
    print(f"File: {input_file_path}")
    
    if not rows_with_404:
        print(f"\n{SYMBOLS['success']} No 404 URLs found - all URLs are working!")
        return
    
    # Handle different output formats
    if output_format == 'console':
        _display_404_console(rows_with_404)
    elif output_format in ['csv', 'xlsx']:
        output_file = _export_404_data(input_file_path, rows_with_404, output_format, df)
        print(f"\n{SYMBOLS['success']} 404 rows exported to: {output_file}")
        _display_404_console(rows_with_404, show_details=False)  # Show summary only
    
    print(f"\n{'='*60}")

def _display_404_console(rows_with_404, show_details=True):
    """Display 404 URLs in console format"""
    if show_details:
        print(f"\n{SYMBOLS['warning']} ROWS WITH 404 URLs:")
        print(f"{'Row':<8} {'Column':<20} {'Error':<15} {'URL':<45}")
        print("-" * 88)
        
        for row_data in rows_with_404:
            for url_data in row_data['404_urls']:
                url_display = url_data['url'][:40] + "..." if len(url_data['url']) > 40 else url_data['url']
                error_display = url_data.get('error', 'HTTP 404')[:12] + "..." if len(url_data.get('error', 'HTTP 404')) > 12 else url_data.get('error', 'HTTP 404')
                print(f"{row_data['row_number']:<8} {url_data['column']:<20} {error_display:<15} {url_display:<45}")
    
    print(f"\n{SYMBOLS['folder']} Summary by column:")
    column_counts = {}
    for row_data in rows_with_404:
        for url_data in row_data['404_urls']:
            column = url_data['column']
            column_counts[column] = column_counts.get(column, 0) + 1
    
    for column, count in sorted(column_counts.items()):
        print(f"  {column}: {count} 404 URLs")

def _export_404_data(input_file_path, rows_with_404, output_format, original_df):
    """Export 404 URLs data to CSV or XLSX with original row data"""
    # Get row indices that have 404s
    rows_with_404_indices = [row_data['row_index'] for row_data in rows_with_404]
    
    # Filter original dataframe to only rows with 404s
    filtered_df = original_df.loc[rows_with_404_indices].copy()
    
    # Create a mapping of 404 URLs for easy lookup
    url_404_map = {}
    for row_data in rows_with_404:
        row_index = row_data['row_index']
        for url_data in row_data['404_urls']:
            url_404_map[(row_index, url_data['column'])] = url_data['error']
    
    # Generate output filename
    input_name = Path(input_file_path).stem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"{input_name}_404_rows_{timestamp}.{output_format}"
    output_path = OUTPUT_DIR / output_filename
    
    # Export data
    try:
        if output_format == 'csv':
            # For CSV, add a note column to indicate 404 URLs
            csv_df = filtered_df.copy()
            csv_df['404_URLs_Found'] = ''
            
            for index, row in csv_df.iterrows():
                row_404_info = []
                for row_data in rows_with_404:
                    if row_data['row_index'] == index:
                        for url_data in row_data['404_urls']:
                            row_404_info.append(f"{url_data['column']}: {url_data['error']}")
                        break
                csv_df.at[index, '404_URLs_Found'] = '; '.join(row_404_info)
            
            csv_df.to_csv(output_path, index=False, encoding='utf-8')
            
        elif output_format == 'xlsx':
            from openpyxl.styles import Font, PatternFill
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write main data
                filtered_df.to_excel(writer, sheet_name='Rows_with_404s', index=False)
                
                # Get the workbook and worksheet for styling
                workbook = writer.book
                worksheet = writer.sheets['Rows_with_404s']
                
                # Define red text format for 404 cells
                red_font = Font(color='FF0000', bold=True)
                red_fill = PatternFill(start_color='FFEEEE', end_color='FFEEEE', fill_type='solid')
                
                # Apply formatting to 404 cells
                for row_idx, (df_index, row) in enumerate(filtered_df.iterrows(), start=2):  # start=2 for Excel header
                    for col_name in filtered_df.columns:
                        if (df_index, col_name) in url_404_map:
                            # Find Excel column index
                            col_idx = filtered_df.columns.get_loc(col_name) + 1  # +1 for Excel 1-indexing
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.font = red_font
                            cell.fill = red_fill
                            # Add comment with error details
                            from openpyxl.comments import Comment
                            error_msg = url_404_map[(df_index, col_name)]
                            comment = Comment(f"404 ERROR: {error_msg}", "URL_Checker")
                            cell.comment = comment
                
                # Add summary sheet with 404 details
                summary_data = []
                for row_data in rows_with_404:
                    for url_data in row_data['404_urls']:
                        summary_data.append({
                            'Row_Number': row_data['row_number'],
                            'Column_Name': url_data['column'],
                            'URL': url_data['url'],
                            'Error': url_data.get('error', 'HTTP 404'),
                            'Source_File': Path(input_file_path).name
                        })
                
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='404_Details', index=False)
                    
                    # Add column count summary
                    column_counts = summary_df['Column_Name'].value_counts().reset_index()
                    column_counts.columns = ['Column', '404_Count']
                    column_counts.to_excel(writer, sheet_name='404_Summary', index=False)
        
        logging.info(f"404 rows exported to: {output_path}")
        return output_path
        
    except Exception as e:
        print(f"Error exporting data: {e}")
        return None

def process_single_url_command(url, folder_name=None):
    """Process a single URL command and return results"""
    log_file = setup_logging(f"single_url_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    
    print(f"{SYMBOLS['upload']} Processing single URL: {url}")
    print(f"   Cloudinary folder: {folder_name or CLOUDINARY_FOLDER}")
    print(f"   Log file: {log_file}")
    
    session = create_robust_session()
    
    try:
        # Download image
        image_data, download_error = download_image(url, session)
        
        if download_error:
            print(f"\n{SYMBOLS['failed']} Download failed: {download_error}")
            return
        
        # Generate filename
        filename = get_url_filename(url) or "single_image"
        
        # Upload to Cloudinary
        result = upload_to_cloudinary(image_data, filename, folder_name)
        
        # Print result
        print(f"\n{'='*60}")
        print(f"{SYMBOLS['upload']} SINGLE URL PROCESSING RESULT")
        print(f"{'='*60}")
        print(f"Original URL: {url}")
        print(f"Status: {result['status']}")
        
        if result['status'] == 'success':
            print(f"{SYMBOLS['success']} Cloudinary URL: {result['cloudinary_url']}")
            print(f"Public ID: {result['public_id']}")
            if result.get('compressed'):
                print(f"Image compressed to {result.get('size_mb', 0):.2f} MB")
        else:
            print(f"{SYMBOLS['failed']} Error: {result.get('error', 'Unknown error')}")
        
        print(f"{'='*60}")
        
    finally:
        session.close()

def main():
    """Main function to handle command line arguments"""
    parser = argparse.ArgumentParser(
        description="Convert CDN URLs to Cloudinary URLs with enterprise features",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process a CSV file
  python scripts/url_tocloudinary.py input.csv
  
  # Process with custom settings
  python scripts/url_tocloudinary.py input.xlsx output.xlsx --workers 8 --folder "my_folder"
  
  # Process single URL
  python scripts/url_tocloudinary.py --url "https://example.com/image.jpg"
  
  # List 404 URLs in console
  python scripts/url_tocloudinary.py input.csv --list-404
  
  # Export 404 URLs to CSV
  python scripts/url_tocloudinary.py input.csv --list-404 --output-format csv
  
  # Export 404 URLs to Excel
  python scripts/url_tocloudinary.py input.csv --list-404 --output-format xlsx
  
  # Force rescan (ignore cache)
  python scripts/url_tocloudinary.py input.csv --force-rescan
        """
    )
    
    # File or URL input
    parser.add_argument('input_file', nargs='?', help='Input file (CSV, XLS, XLSX)')
    parser.add_argument('output_file', nargs='?', help='Output XLSX file (optional)')
    
    # Single URL mode
    parser.add_argument('--url', help='Process a single image URL')
    
    # Options
    parser.add_argument('--folder', default='url_uploads', help='Cloudinary folder name (default: url_uploads)')
    parser.add_argument('--workers', type=int, help='Number of concurrent workers (default: auto)')
    parser.add_argument('--force-rescan', action='store_true', help='Ignore cache and reprocess all URLs')
    parser.add_argument('--list-404', action='store_true', help='Test all URLs and list those returning 404 errors')
    parser.add_argument('--output-format', choices=['console', 'csv', 'xlsx'], default='xlsx', help='Output format for --list-404 (default: xlsx)')
    
    args = parser.parse_args()
    
    # Single URL mode
    if args.url:
        process_single_url_command(args.url, args.folder)
        return
    
    # File mode
    if not args.input_file:
        parser.print_help()
        return
    
    if not os.path.exists(args.input_file):
        print(f"Error: Input file not found: {args.input_file}")
        return
    
    # List 404 URLs mode
    if args.list_404:
        list_404_urls(args.input_file, args.output_format)
        return
    
    # Process the file
    process_file_to_cloudinary(
        input_file_path=args.input_file,
        output_xlsx_path=args.output_file,
        max_workers=args.workers,
        folder_name=args.folder,
        force_rescan=args.force_rescan
    )

if __name__ == "__main__":
    main()